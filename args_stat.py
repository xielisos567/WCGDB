#!/bin/python3

import os
import sys
import openpyxl
import pandas as pd
from collections import defaultdict

def getAllInfo(all_info):
	########## get all_info file data ############
	all_info_file = open(all_info, encoding='utf-8')
	headers = next(all_info_file).strip().split('\t')
	# get data index
	cdc_index = headers.index('collection_date_curated')
	hdis_index = headers.index('Host_Host_disease_Isolation_source')
	continent_index = headers.index('Continent')
	genome_index = headers.index('Genomes')
	# get all infile data
	cdc_data = []
	hdis_data = []
	continent_data = []
	genome_info = {}
	years = {'before 2000':[0,1999], '2000-2004':[2000, 2004],
			'2005-2009':[2005,2009], '2010-2014':[2010,2014],
			'2015-2019':[2015,2019], '2020-2023':[2020,2023]}
	for line in all_info_file:
		line_info = line.strip().split('\t')
		cdc_year = line_info[cdc_index]
		try:
			cdc_year = int(cdc_year)
			for key in years:
				start, end = years[key]
				if cdc_year>=start and cdc_year<=end:
					cdc_value = key
					break
		except:
			cdc_value = cdc_year
		hdis_value = line_info[hdis_index]
		continent_value = line_info[continent_index]
		genome_info[line_info[genome_index]] = [cdc_value, hdis_value, continent_value]
		#cdc_data.append(cdc_value) if cdc_value not in cdc_data else ''
		cdc_data = ['before 2000', '2000-2004', '2005-2009', '2010-2014', '2015-2019', '2020-2023']
		hdis_data.append(hdis_value) if (hdis_value not in hdis_data and hdis_value!='Not collected') else ''
		continent_data.append(continent_value) if (continent_value not in continent_data and continent_value!='Not collected') else ''
	return genome_info, cdc_data, sorted(hdis_data), sorted(continent_data)

def statArgs(argfile, genomeinfo, cdclist, hdislist, continentlist, out):
	'''
	'''
	args_file = open(argfile)
	args_data = defaultdict(dict)
	for line in args_file:
		line_info = line.strip().split('\t')
		genome, argid = line_info[:2]
		genome_type = line_info[-1]
		args_data[';'.join([argid, genome_type])][genome] = genome_info[genome]
	args_file.close()
	out1 = open(f'{out}_extract.txt', 'w')
	out1.write('ARGs\tType\tGenome\tcollection_date_curated\tHost_Host_disease_Isolation_source\tContinent\n')
	all_data = defaultdict(list)
	header_list = ['Genomes']
	cols = [('ARGs',''), ('Type', ''), ('Genomes', '')]
	header_list.extend(cdclist)
	for item in cdclist:
		cols.append(('collection_date_curated', item))
	header_list.extend(hdislist)
	for item in hdislist:
		cols.append(('Host_Host_disease_Isolation_source', item))
	header_list.extend(continentlist)
	for item in continentlist:
		cols.append(('Continent', item))
	for key in args_data:
		argsid, argstype = key.split(';')
		all_data['ARGs'].append(argsid)
		all_data['Types'].append(argstype)
		genomes_data = args_data[key]
		item_count = defaultdict(int)
		for genome in genomes_data:
			item_count['Genomes'] += 1
			genome_o = '\t'.join(genomes_data[genome])
			out1.write(f'{argsid}\t{argstype}\t{genome}\t{genome_o}\n')
			for item in genomes_data[genome]:
				item_count[item] += 1
		for head in header_list:
			all_data[head].append(item_count[head])
	#print (all_data)
	all_data = pd.DataFrame(all_data)
	all_data.columns = pd.MultiIndex.from_tuples(cols)
	all_data.to_csv(f'{outprefix}_stat.xls', sep='\t', index=False)
	writer = pd.ExcelWriter(f'{out}.xlsx')
	all_data.to_excel(writer)
	writer.close()
	excel_file = openpyxl.load_workbook(f'{out}.xlsx')
	excel_data = excel_file['Sheet1']
	excel_data.delete_rows(3)
	excel_data.merge_cells('D1:D2')
	excel_data.merge_cells('B1:B2')
	excel_data.merge_cells('C1:C2')
	excel_file.save(f'{out}.xlsx')

if __name__ == '__main__':
	if len(sys.argv) != 4:
		print (f'Usage:\n    python3 {sys.argv[0]} all_info[in] args_data[in] outprefix')
		sys.exit()
	all_info, argfile, outprefix = sys.argv[1:]
	# get genome info
	genome_info, cdc_info, hdis_info, continent_info = getAllInfo(all_info)
	# stat infile
	statArgs(argfile, genome_info, cdc_info, hdis_info, continent_info, outprefix)